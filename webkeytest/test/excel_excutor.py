import openpyxl
from openpyxl.styles import PatternFill, Font

from webkeytest.keywords.key_words import KeyWords


# excel_case为excel测试用例路径
class ExcelExcutor:
    def excel_case_excute(self, excel_case):
        # 获取excel工作簿
        excel = openpyxl.load_workbook(excel_case)
        sheet = excel['Sheet3']
        args = {}

        for value in sheet.values:
            args['name'] = value[2]
            args['value'] = value[3]
            args['text'] = value[4]
            args['fact_text'] = value[6]
            print(value)
            if type(value[0]) is int:
                if value[1] == 'open_browser':
                    print(value[4])
                    wb = KeyWords(value[4], '../log/info.log')
                elif 'assert_text' in value[1]:
                    status = getattr(wb, value[1])(**args)
                    print(status)
                    if status:
                        sheet.cell(row=value[0] + 1, column=8).value = 'Pass'
                        sheet.cell(row=value[0] + 1, column=8).fill = PatternFill('solid', fgColor='AACF91')
                        sheet.cell(row=value[0] + 1, column=8).font = Font(bold=True)
                    else:
                        sheet.cell(row=value[0] + 1, column=8).value = 'Failed'
                        sheet.cell(row=value[0] + 1, column=8).fill = PatternFill('solid', fgColor='FF0000')
                        sheet.cell(row=value[0] + 1, column=8).font = Font(bold=True)
                    excel.save(excel_case)
                # 判断用例的方法是否存在
                elif hasattr(wb, value[1]):
                    # print(hasattr(wb,value[1]))
                    getattr(wb, value[1])(**args)
                else:
                    print('没找事件{}'.format(value[1]))

        excel.close()

if __name__ == '__main__':
    ee = ExcelExcutor()
    ee.excel_case_excute('../case_excels/test_data.xlsx')
