'''
@Project ：auto_test 
@File ：excel_excutor.py
@Date ：2021/3/30 9:37 
'''
import openpyxl
from openpyxl.styles import PatternFill, Font

from apikeytest.keywords.key_words import ApiKeys
from common.log import Logging


class ApiExcelParse:


    def excelparse(self, excel_path, sheet_name):
        '''
        :param excel_path: 执行接口测试用例excel路径
        :param sheet_name: 指定excel用例工作簿名称
        :return: 无返回
        '''
        # 获取excel表对象
        excel = openpyxl.load_workbook(excel_path)
        # 获取excel表的操作工作簿sheet1
        sheet = excel[sheet_name]
        # 工作簿sheet1里的数据
        for value in sheet.values:
            arg = {}
            print(value)
            if type(value[0]) is int:
                # 获取关键字方法
                ak = ApiKeys()
                # 造数据
                arg['url'] = value[1] + value[2]
                if value[4] is not None:
                    arg['headers'] = eval(value[4])
                if value[5] is not None:
                    arg[value[6]] = eval(value[5])
                method = value[3]
                expect_field = value[7]
                expect_value = value[8]
                # 通过反射方法请求接口数据
                response = getattr(ak, method)(**arg)
                # print(response)
                # log.info('返回结果：{}'.format('1'))
                # 响应内容中查找是否存在期待值，异常返回False,不存在返回对应信息
                text = ak.get_text(response.text, expect_field)
                # 接口用例执行成功Pass，失败则为False
                if text == expect_value:
                    sheet.cell(row=value[0] + 1, column=10).value = 'Pass'
                    sheet.cell(row=value[0] + 1, column=10).fill = PatternFill('solid', fgColor='AACF91')
                    sheet.cell(row=value[0] + 1, column=10).font = Font(bold=True)
                else:
                    sheet.cell(row=value[0] + 1, column=10).value = 'Failed'
                    sheet.cell(row=value[0] + 1, column=10).fill = PatternFill('solid', fgColor='FF0000')
                    sheet.cell(row=value[0] + 1, column=10).font = Font(bold=True)
                excel.save(excel_path)
        excel.close()


if __name__ == '__main__':
    aep = ApiExcelParse()
    aep.excelparse('../data/login_cases.xlsx', 'Sheet1')
